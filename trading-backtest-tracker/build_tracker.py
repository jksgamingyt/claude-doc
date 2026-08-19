import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Backtest Tracker"

# ---- Theme colors ----
NAVY = "1B2A4A"
TEAL = "0F766E"
GOLD = "C9A227"
LIGHT_ROW = "F4F6F8"
WHITE = "FFFFFF"
DARK_TEXT = "1B2A4A"
BORDER_COLOR = "D9DEE4"

FONT_NAME = "Calibri"

thin = Side(style="thin", color=BORDER_COLOR)
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_row_height(row, height):
    ws.row_dimensions[row].height = height

# Column widths
ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20

# ---- Title banner ----
ws.merge_cells("A1:D1")
c = ws["A1"]
c.value = "TRADING STRATEGY BACKTEST TRACKER"
c.font = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
set_row_height(1, 34)

ws.merge_cells("A2:D2")
c = ws["A2"]
c.value = "Confluence Checklist  +  50-Trade Backtest Log"
c.font = Font(name=FONT_NAME, size=11, italic=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
set_row_height(2, 22)

# spacer row 3
set_row_height(3, 8)

# ---- Confluence section header ----
ws.merge_cells("A4:D4")
c = ws["A4"]
c.value = "CONFLUENCE CHECKLIST"
c.font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=TEAL)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
set_row_height(4, 26)

ws.merge_cells("A5:D5")
c = ws["A5"]
c.value = "List the confluence factors your strategy requires before you take a trade."
c.font = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
set_row_height(5, 18)

# Confluence table headers
conf_headers_row = 6
headers = ["#", "Confluence Factor", "Description / Notes", ""]
ws.merge_cells(start_row=conf_headers_row, start_column=3, end_row=conf_headers_row, end_column=4)
for idx, h in enumerate(["#", "Confluence Factor", "Description / Notes"], start=1):
    cell = ws.cell(row=conf_headers_row, column=idx, value=h)
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=GOLD)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_all
set_row_height(conf_headers_row, 20)

# 10 blank confluence rows
conf_start = conf_headers_row + 1
conf_count = 10
for i in range(conf_count):
    r = conf_start + i
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    num_cell = ws.cell(row=r, column=1, value=i + 1)
    num_cell.alignment = Alignment(horizontal="center", vertical="center")
    fill_color = WHITE if i % 2 == 0 else LIGHT_ROW
    for col in range(1, 5):
        cell = ws.cell(row=r, column=col)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.border = border_all
        cell.font = Font(name=FONT_NAME, size=10, color=DARK_TEXT)
    set_row_height(r, 18)

conf_end = conf_start + conf_count - 1

# spacer
spacer1 = conf_end + 1
set_row_height(spacer1, 10)

# ---- Backtest log section header ----
log_header_row = spacer1 + 1
ws.merge_cells(start_row=log_header_row, start_column=1, end_row=log_header_row, end_column=4)
c = ws.cell(row=log_header_row, column=1, value="BACKTEST TRADE LOG")
c.font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=TEAL)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
set_row_height(log_header_row, 26)

note_row = log_header_row + 1
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
c = ws.cell(row=note_row, column=1, value="Log each backtested trade below. Use the dropdown in the Result column.")
c.font = Font(name=FONT_NAME, size=10, italic=True, color="6B7280")
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
set_row_height(note_row, 18)

table_header_row = note_row + 1
table_headers = ["Stock", "Reasoning for Trade Entry", "Timeframe(s) Used", "Result (Success / Fail)"]
for idx, h in enumerate(table_headers, start=1):
    cell = ws.cell(row=table_header_row, column=idx, value=h)
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_all
set_row_height(table_header_row, 30)

data_start = table_header_row + 1
NUM_ROWS = 50
data_end = data_start + NUM_ROWS - 1

# Data validation dropdown for Result column
dv = DataValidation(type="list", formula1='"Success,Fail"', allow_blank=True, showDropDown=False)
ws.add_data_validation(dv)

for i in range(NUM_ROWS):
    r = data_start + i
    fill_color = WHITE if i % 2 == 0 else LIGHT_ROW
    for col in range(1, 5):
        cell = ws.cell(row=r, column=col)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.border = border_all
        cell.font = Font(name=FONT_NAME, size=10, color=DARK_TEXT)
        cell.alignment = Alignment(horizontal="center" if col in (1, 3, 4) else "left", vertical="center", wrap_text=(col == 2))
    dv.add(ws.cell(row=r, column=4))
    set_row_height(r, 18)

# Rows ship blank as a template; these rules color the Result cell once it's filled in.
green_fill = PatternFill("solid", fgColor="D1FAE5")
green_font = Font(name=FONT_NAME, size=10, color="065F46", bold=True)
red_fill = PatternFill("solid", fgColor="FEE2E2")
red_font = Font(name=FONT_NAME, size=10, color="991B1B", bold=True)

result_range = f"D{data_start}:D{data_end}"
ws.conditional_formatting.add(
    result_range,
    CellIsRule(operator="equal", formula=['"Success"'], fill=green_fill, font=green_font),
)
ws.conditional_formatting.add(
    result_range,
    CellIsRule(operator="equal", formula=['"Fail"'], fill=red_fill, font=red_font),
)

# spacer
spacer2 = data_end + 1
set_row_height(spacer2, 10)

# ---- Summary section ----
summary_header_row = spacer2 + 1
ws.merge_cells(start_row=summary_header_row, start_column=1, end_row=summary_header_row, end_column=4)
c = ws.cell(row=summary_header_row, column=1, value="SUMMARY STATS")
c.font = Font(name=FONT_NAME, size=13, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=GOLD)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
set_row_height(summary_header_row, 26)

labels_row = summary_header_row + 1
stats = [
    ("Total Trades Logged", f"=COUNTIF(D{data_start}:D{data_end},\"Success\")+COUNTIF(D{data_start}:D{data_end},\"Fail\")"),
    ("Wins", f"=COUNTIF(D{data_start}:D{data_end},\"Success\")"),
    ("Losses", f"=COUNTIF(D{data_start}:D{data_end},\"Fail\")"),
    ("Win Rate", f"=IFERROR(C{labels_row+1}/C{labels_row}, 0)"),
]

# lay out as label (A) merged B, value in C merged D, one stat per row
for i, (label, formula) in enumerate(stats):
    r = labels_row + i
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    lbl_cell = ws.cell(row=r, column=1, value=label)
    lbl_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK_TEXT)
    lbl_cell.fill = PatternFill("solid", fgColor=LIGHT_ROW)
    lbl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl_cell.border = border_all
    ws.cell(row=r, column=2).border = border_all
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=LIGHT_ROW)

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    val_cell = ws.cell(row=r, column=3, value=formula)
    val_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=DARK_TEXT)
    val_cell.fill = PatternFill("solid", fgColor=WHITE)
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    val_cell.border = border_all
    ws.cell(row=r, column=4).border = border_all
    if label == "Win Rate":
        val_cell.number_format = "0.0%"
    set_row_height(r, 18)

# Freeze panes so the trade log header stays visible while scrolling the 50 rows
ws.freeze_panes = f"A{table_header_row + 1}"

# Gridlines off for a cleaner themed look
ws.sheet_view.showGridLines = False

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Trading_Backtest_Tracker.xlsx")
wb.save(OUT)
print("saved", OUT)
print("data_start", data_start, "data_end", data_end, "labels_row", labels_row)
